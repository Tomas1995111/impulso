import os
import time
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import pyautogui
from openpyxl import load_workbook

def generar_resumen_alertas():
    archivo = "C:/Users/Tomas/OneDrive/Escritorio/impulso_wsp_bot/registros/alertas.xlsx"
    df = pd.read_excel(archivo)
    df["Fecha Alerta"] = pd.to_datetime(df["Fecha Alerta"], errors='coerce')

    hoy = datetime.now()

    for i, row in df.iterrows():
        ticker = row["Ticker"]
        fecha_alerta = row["Fecha Alerta"]
        precio_alerta = row["Precio Alerta"]
        stop_loss = row["Stop Loss"]

        if pd.isna(fecha_alerta):
            continue

        # Convertir fechas a strings en formato 'YYYY-MM-DD' para yf.download
        start_str = fecha_alerta.strftime("%Y-%m-%d")
        end_str = (hoy + timedelta(days=1)).strftime("%Y-%m-%d")

        try:
            data = yf.download(ticker, start=start_str, end=end_str)
        except Exception as e:
            print(f"Error al descargar datos de {ticker}: {e}")
            continue

        if data.empty or "Low" not in data or "Close" not in data:
            continue

        # Convertir a float para evitar errores de comparación con Series
        try:
            minimo_val = float(data["Low"].min())
            actual_val = float(data["Close"].iloc[-1])
            stop_loss_val = float(stop_loss)
            precio_alerta_val = float(precio_alerta)
        except Exception as e:
            print(f"Error en conversión de valores numéricos para {ticker} en fila {i+2}: {e}")
            continue

        df.at[i, "Minimo"] = minimo_val
        df.at[i, "Precio Actual"] = actual_val

        if minimo_val < stop_loss_val:
            df.at[i, "Analisis"] = "Saltó el SL"
        elif actual_val < stop_loss_val:
            df.at[i, "Analisis"] = "SL"
        else:
            variacion = (actual_val - precio_alerta_val) / precio_alerta_val
            df.at[i, "Analisis"] = round(variacion, 4)

        # Formatea la fecha solo si es datetime
        if isinstance(fecha_alerta, (pd.Timestamp, datetime)):
            df.at[i, "Fecha Alerta"] = fecha_alerta.strftime('%d/%m/%Y')

    actualizar_excel_sin_perder_formato(df, archivo)

def actualizar_excel_sin_perder_formato(df_actualizado, archivo):
    wb = load_workbook(archivo)
    ws = wb.active

    for i, row in df_actualizado.iterrows():
        fila = i + 2  # Asume encabezados en fila 1
        ws[f"E{fila}"] = row.get("Minimo", "")
        ws[f"F{fila}"] = row.get("Precio Actual", "")
        ws[f"G{fila}"] = row.get("Analisis", "")
    wb.save(archivo)

def seleccionar_y_copiar_excel():
    ruta_excel = r"C:\Users\Tomas\OneDrive\Escritorio\impulso_wsp_bot\registros\alertas.xlsx"
    os.startfile(ruta_excel)
    time.sleep(5)

    pyautogui.hotkey('ctrl', 'up')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'left')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'shift', 'space')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

if __name__ == "__main__":
    generar_resumen_alertas()
    seleccionar_y_copiar_excel()
