# mensajesAlertaCompra.py
import yfinance as yf
import random
import math
import datetime
import openpyxl
import time

tickers_top = [
    'NOW', 'SHW', 'COST', 'AZO', 'SNPS', 
    'META', 'LMT', 'CAT', 'TMO', 'UNH', 
    'DE', 'ADSK', 'IBM', 'JPM', 'AAPL', 
    'UNP', 'HD', 'BLK', 'PNC', 'FDX', 
    'NSC', 'AMZN', 'BRK-B', 'TMUS', 'CRM', 
    'MAR', 'RSG', 'EXPE', 'AXP', 'QCOM', 
    'LOW', 'GE', 'CVX', 'RL', 'VST', 'LIN', 
    'CMI', 'ACN', 'MCD', 'MSFT', 'DIS', 
    'JNJ', 'AMGN', 'HON', 'PG', 'MMM', 
    'BA', 'NVDA', 'KO', 'V', 'WMT', 
    'VZ', 'GS', 'NKE', 'CSCO', 'MRK', 
    'NFLX', 'ASML', 'REGN', 'KLAC', 'BKNG', 
    'MELI', 'MDB', 'MSTR', 'ZS', 'AMD', 
    'AVGO', 'GILD', 'TXN', 'TSLA', 'GOOG', 
    'ROST', 'TTWO', 'WDAY', 'PLTR', 'CEG', 
    'MU', 'LLY', 'MCK', 'GOOGL', 'TSM', 
    'MA', 'ORCL', 'XOM', 'SAP', 'BAC', 
    'ABBV', 'SPY', 'QQQ', 'DIA', 'IWM', 
    'VTI', 'VEA', 'VWO', 'TLT', 'GLD', 
    'XLF', 'XLE', 'XLV', 'XLK', 'XLY', 
    'XLU', 'INTC', 'PEP', 'UPS', 'ADBE', 
    'MDT', 'PFE', 'BABA', 'SBUX', 'CSX'
]

def guardar_en_excel(fecha, ticker, precio, stop_loss, ruta_excel):
    try:
        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb.active

        # Buscar primera fila vacía
        fila = 2
        while ws.cell(row=fila, column=1).value is not None:
            fila += 1

        # Guardar los datos
        ws.cell(row=fila, column=1).value = fecha
        ws.cell(row=fila, column=2).value = ticker
        ws.cell(row=fila, column=3).value = precio
        ws.cell(row=fila, column=4).value = stop_loss

        wb.save(ruta_excel)
        print(f"Datos guardados en fila {fila}")
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")

def obtener_datos_accion(ticker):
    accion = yf.Ticker(ticker)
    info = accion.info

    datos = {
        "ticker": ticker,
        "nombre": info.get("shortName"),
        "precio_actual": info.get("regularMarketPrice"),
        "variacion_pct": info.get("regularMarketChangePercent"),
        "max_dia": info.get("dayHigh"),
        "min_dia": info.get("dayLow"),
        "max_historico": accion.history(period="max")["High"].max() if not accion.history(period="max").empty else None,
        "capitalizacion_bursatil": info.get("marketCap"),
        "pe_ratio": info.get("trailingPE"),
        "rendimiento_dividendos": info.get("dividendYield"),
        "sector": info.get("sector"),
        "recomendacion": info.get("recommendationKey"),
    }

    return datos

def generar_alerta(datos):
    precio_actual = datos["precio_actual"]
    max_historico = datos["max_historico"]
    recomendacion = datos["recomendacion"]

    # Si no hay recomendacion, considerarla como "buy"
    if recomendacion is None or recomendacion.lower() == "none":
        recomendacion = "buy"

    # Aceptar "buy", "strong_buy" y "strongBuy" (por precaución)
    if recomendacion.lower() in ["buy", "strong_buy", "strongbuy"] and precio_actual < 0.8 * max_historico:
        PE = math.floor(precio_actual)
        SL_pct = -random.uniform(6, 14) / 100
        SL = math.floor(PE * (1 + SL_pct))
        R = abs(SL_pct)

        TP1 = math.floor(PE * (1 + R * 0.9))
        TP2 = math.floor(PE * (1 + R * 1.8))
        TP3 = math.floor(PE * (1 + R * 2.6))

        mensaje = f"""
📢 *ALERTA ANÁLISIS* // ESPECULATIVO 
👉🏼 Perfil Agresivo❗
•Ticker: *{datos['ticker']}* ({datos['nombre']}) 🇦🇷🇺🇸
•Zona de compra: {PE} USD
⛔ STOP LOSS = *{round(SL_pct * 100)}%*
✅ DESARMES: ( {TP1} USD / {TP2} USD / {TP3} USD )
- - - - - - - - - - - - - - - - - - - - - - 
Recuerde operar bajo su propio riesgo y en la justa y considerada proporción de su cartera. (la misma no configura ninguna recomendación)
"""
         # Guardar en Excel si hay alerta
        fecha_actual = datetime.date.today().strftime("%Y-%m-%d")
        ruta_excel = "C:\\Users\\Tomas\\OneDrive\\Escritorio\\impulso_wsp_bot\\registros\\alertas.xlsx"
        guardar_en_excel(fecha_actual, datos['ticker'], precio_actual, SL, ruta_excel)

        return mensaje.strip()
    else:
        return None

def generar_alerta_aleatoria():
    """Busca en la lista tickers_top alguna acción que cumpla y devuelve mensaje o None."""

    tickers_restantes = tickers_top.copy()
    random.shuffle(tickers_restantes)

    for ticker in tickers_restantes:
        try:
            datos = obtener_datos_accion(ticker)
            mensaje = generar_alerta(datos)
            if mensaje:
                return mensaje
            time.sleep(1)  # para no saturar requests
        except Exception as e:
            print(f"Error con {ticker}: {e}")
    return None

# Si quieres probar el script directamente
if __name__ == "__main__":
    alerta = generar_alerta_aleatoria()
    if alerta:
        print(alerta)
    else:
        print("No hay alertas en este momento.")
