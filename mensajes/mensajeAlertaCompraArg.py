# mensajeAlertaCompraArg.py
import yfinance as yf
import random
import math
import datetime
import openpyxl
import time

tickers_arg = [
    "GGAL.BA", "YPFD.BA", "BMA.BA", "BBAR.BA", "PAMP.BA",
    "TGSU2.BA", "TXAR.BA", "SUPV.BA", "COME.BA", "BYMA.BA",
    "CEPU.BA", "ALUA.BA", "TRAN.BA", "LOMA.BA", "EDN.BA",
    "VALO.BA", "METR.BA", "IRSA.BA", "TECO2.BA", "TGNO4.BA",
    "CRES.BA", "MIRG.BA", "BOLT.BA", "AUSO.BA", "SAMI.BA", 
    "MOLI.BA", "RICH.BA", "LEDE.BA", "CVH.BA", "BPAT.BA", 
    "DGCU2.BA", "BHIP.BA", "CELU.BA", "AGRO.BA", "PATA.BA",
    "CECO2.BA", "A3.BA", "GRIM.BA", "MORI.BA", "HARG.BA",
    "GBAN.BA", "CGPA2.BA",    
]

def guardar_en_excel(fecha, ticker, precio, stop_loss, ruta_excel):
    try:
        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb.active

        # Buscar primera fila vacía
        fila = 2
        while ws.cell(row=fila, column=1).value is not None:
            fila += 1

        # Guardar los datos
        ws.cell(row=fila, column=1).value = fecha
        ws.cell(row=fila, column=2).value = ticker
        ws.cell(row=fila, column=3).value = precio
        ws.cell(row=fila, column=4).value = stop_loss

        wb.save(ruta_excel)
        print(f"Datos guardados en fila {fila}")
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")

def obtener_datos_accion(ticker):
    accion = yf.Ticker(ticker)
    info = accion.info

    datos = {
        "ticker": ticker,
        "nombre": info.get("shortName"),
        "precio_actual": info.get("regularMarketPrice"),
        "variacion_pct": info.get("regularMarketChangePercent"),
        "max_dia": info.get("dayHigh"),
        "min_dia": info.get("dayLow"),
        "max_historico": accion.history(period="max")["High"].max() if not accion.history(period="max").empty else None,
        "capitalizacion_bursatil": info.get("marketCap"),
        "pe_ratio": info.get("trailingPE"),
        "rendimiento_dividendos": info.get("dividendYield"),
        "sector": info.get("sector"),
        "recomendacion": info.get("recommendationKey"),
    }

    return datos

def generar_alerta(datos):
    precio_actual = datos["precio_actual"]
    max_historico = datos["max_historico"]
    recomendacion = datos["recomendacion"]

    # Si no hay recomendacion, considerarla como "buy"
    if recomendacion is None or recomendacion.lower() == "none":
        recomendacion = "buy"

    # Aceptar "buy", "strong_buy" y "strongBuy" (por precaución)
    if recomendacion.lower() in ["buy", "strong_buy", "strongbuy"] and precio_actual < 0.8 * max_historico:
        PE = math.floor(precio_actual)
        SL_pct = -random.uniform(6, 14) / 100
        SL = math.floor(PE * (1 + SL_pct))
        R = abs(SL_pct)

        TP1 = math.floor(PE * (1 + R * 0.9))
        TP2 = math.floor(PE * (1 + R * 1.8))
        TP3 = math.floor(PE * (1 + R * 2.6))

        mensaje = f"""
📢 *ALERTA ANÁLISIS* // ESPECULATIVO 
👉🏼 Perfil Agresivo❗
•Ticker: *{datos['ticker']}* ({datos['nombre']}) 🇦🇷
•Zona de compra: {PE} ARS
⛔ STOP LOSS = *{round(SL_pct * 100)}%*
✅ DESARMES: ( {TP1} ARS / {TP2} ARS / {TP3} ARS )
- - - - - - - - - - - - - - - - - - - - - - 
Recuerde operar bajo su propio riesgo y en la justa y considerada proporción de su cartera. (la misma no configura ninguna recomendación)
"""
         # Guardar en Excel si hay alerta
        fecha_actual = datetime.date.today().strftime("%Y-%m-%d")
        ruta_excel = "C:\\Users\\Tomas\\OneDrive\\Escritorio\\impulso_wsp_bot\\registros\\alertas.xlsx"
        guardar_en_excel(fecha_actual, datos['ticker'], precio_actual, SL, ruta_excel)

        return mensaje.strip()
    else:
        return None

def generar_alerta_aleatoria_arg():
    """Busca en la lista tickers_arg alguna acción que cumpla y devuelve mensaje o None."""

    tickers_restantes = tickers_arg.copy()
    random.shuffle(tickers_restantes)

    for ticker in tickers_restantes:
        try:
            datos = obtener_datos_accion(ticker)
            mensaje = generar_alerta(datos)
            if mensaje:
                return mensaje
            time.sleep(1)  # para no saturar requests
        except Exception as e:
            print(f"Error con {ticker}: {e}")
    return None

# Si quieres probar el script directamente
if __name__ == "__main__":
    alerta = generar_alerta_aleatoria_arg()
    if alerta:
        print(alerta)
    else:
        print("No hay alertas en este momento.")
